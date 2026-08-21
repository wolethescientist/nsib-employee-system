"""Extract the "Training Organisations" sheet into data/training-organisations.json.

The sheet is the bureau's directory of training schools: a numbered list in
columns F-K with name, website, email, telephone and address. Addresses spill
down over several rows beneath their entry, so a row that has no number in
column F is a continuation of the entry above it.

Run:  python scripts/extract-organisations.py
"""
import json
import os
import re

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WORKBOOK = os.path.join(ROOT, "AIA Training Program Management.xlsx")
OUT = os.path.join(ROOT, "data", "training-organisations.json")

SHEET = "Training Organisations"
FIRST_ROW = 6
COL = {"no": 6, "name": 7, "website": 8, "email": 9, "phone": 10, "address": 11}


def clean(value):
    """The sheet is full of non-breaking and narrow no-break spaces."""
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").replace(" ", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text or None


def tidy_url(value):
    if not value:
        return None
    url = value.replace("https:// ", "https://").replace("http:// ", "http://")
    return url if url.startswith("http") else f"https://{url}"


def main():
    sheet = openpyxl.load_workbook(WORKBOOK, data_only=True)[SHEET]
    organisations = []

    for row in range(FIRST_ROW, sheet.max_row + 1):
        cell = lambda key: clean(sheet.cell(row=row, column=COL[key]).value)
        number = cell("no")

        if number:
            organisations.append(
                {
                    "serial": int(re.sub(r"\D", "", number) or 0),
                    "name": cell("name"),
                    "website": tidy_url(cell("website")),
                    "emails": [],
                    "phones": [],
                    "address": [],
                    "contact": None,
                }
            )
        if not organisations:
            continue

        # The entry's own row and the rows beneath it are read the same way: the
        # address spills down over several rows, and the email column also
        # carries the odd second website and the odd named contact — route those
        # by shape rather than filing a URL as an email address.
        entry = organisations[-1]
        for key, field in (("email", "emails"), ("phone", "phones"), ("address", "address")):
            value = cell(key)
            if not value:
                continue
            if key == "email" and "@" not in value:
                if value.startswith(("http", "www.")):
                    entry["website"] = entry["website"] or tidy_url(value)
                else:
                    entry["contact"] = value
                continue
            entry[field].append(value)

    # Entries with no name are stray formatting, not organisations.
    organisations = [o for o in organisations if o["name"]]
    for entry in organisations:
        entry["address"] = ", ".join(entry.pop("address")) or None
        entry["email"] = ", ".join(entry.pop("emails")) or None
        entry["phone"] = ", ".join(entry.pop("phones")) or None

    with open(OUT, "w", encoding="utf-8") as handle:
        json.dump({"organisations": organisations}, handle, indent=1, ensure_ascii=False)

    print(f"{len(organisations)} training organisations -> {os.path.relpath(OUT, ROOT)}")


if __name__ == "__main__":
    main()
