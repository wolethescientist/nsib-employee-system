"""Extract the annual training plan workbook into data/training-plan.json.

One sheet per year ("2024", "2025", "2026"), laid out as:

    S/N | NAME | COURSE TITLE | INSTITUTION/COUNTRY | DATE | PRIORITY | TRAINING TYPE | COURSE FEE

A name in the NAME column opens a block; the rows under it, with the name column
blank, are the rest of that person's courses for the year.

Run:  python scripts/extract-training-plan.py
"""
import datetime
import json
import os
import re

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WORKBOOK = os.path.join(ROOT, "Training Plan.xlsx")
OUT = os.path.join(ROOT, "data", "training-plan.json")

# A, B, C ... as 1-based column indices.
COL = {"sn": 1, "name": 2, "title": 3, "institution": 4, "date": 5, "priority": 6, "type": 7, "fee": 8}

PRIORITY = {"p1": "P1", "p2": "P2", "p3": "P3", "r": "R", "high": "P1", "medium": "P2", "low": "P3"}

# The sheet's spelling of the training type, normalised to one word per idea.
TYPE_FIX = {
    "specialize": "Specialize",
    "specialized": "Specialize",
    "specialty": "Specialty",
    "specialities": "Specialty",
    "advance": "Advance",
    "advanced": "Advance",
    "recurent": "Recurrent",
    "recurrent": "Recurrent",
    "initial": "Initial",
    "basic": "Basic",
    "additional": "Additional",
    "ojt": "OJT",
    "initial/advance": "Initial",
}

CURRENCY_MARKS = [("$", "USD"), ("£", "GBP"), ("€", "EUR"), ("₦", "NGN"), ("=n=", "NGN"), ("n=", "NGN"), ("ngn", "NGN")]

# Excel's 1900 epoch. A cell holding the bare number 2026 and formatted as a date
# comes back as 1905-07-18, which is 2026 days after the epoch — so anything that
# lands in the first few years of the epoch is a number wearing a date's clothes.
EPOCH = datetime.datetime(1899, 12, 30)
SERIAL_IS_REALLY_A_NUMBER = 3000  # ~1908; no training was planned in 1908.


def clean(value):
    """A cell as trimmed text. Dates that are really numbers come back as numbers."""
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        serial = (value - EPOCH).days
        if serial < SERIAL_IS_REALLY_A_NUMBER:
            return str(serial)
        return value.strftime("%d %b %Y").lstrip("0")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def parse_amount(raw):
    """"$10,340" -> (10340.0, "USD");  "=N= 350,000" -> (350000.0, "NGN");  "-" -> (None, None).

    The sheet quotes dollars, pounds and naira in the same column, so the symbol
    decides the currency and everything that is not a digit is discarded.
    """
    text = clean(raw).lower()
    if not text:
        return None, None
    currency = next((code for mark, code in CURRENCY_MARKS if mark in text), None)
    digits = re.sub(r"[^0-9.]", "", text.replace(",", ""))
    # A stray full stop ("- . -") is not a number; a real figure has a digit in it.
    if not re.search(r"[0-9]", digits):
        return None, None
    try:
        amount = float(digits.rstrip("."))
    except ValueError:
        return None, None
    if amount <= 0:
        return None, None
    return amount, currency or "NGN"


def parse_serial(title):
    """"3) Analysis" -> (3, "Analysis"). Numbering is the person's own list order."""
    match = re.match(r"^\s*\(?(\d{1,2})\s*[).:-]\s*(.*)$", title)
    if match:
        return int(match.group(1)), match.group(2).strip()
    return None, title.strip()


def read_year(sheet):
    """Every course line on one year's sheet, grouped under the name above it."""
    rows = []
    current = None
    counter = 0
    for row in sheet.iter_rows(min_row=1, values_only=True):
        cell = lambda key: clean(row[COL[key] - 1]) if len(row) >= COL[key] else ""

        name = cell("name")
        title_raw = cell("title")

        # The header row, and the banner rows above it.
        if name.upper() in {"NAME", ""} and title_raw.upper() in {"COURSE TITLE", ""}:
            if name.upper() == "NAME":
                current = None
            if not title_raw:
                continue
        if title_raw.upper() == "COURSE TITLE":
            continue

        if name and name.upper() != "NAME":
            current = name
            counter = 0
        if not current or not title_raw:
            continue

        serial, title = parse_serial(title_raw)
        counter += 1
        amount, currency = parse_amount(row[COL["fee"] - 1] if len(row) >= COL["fee"] else None)
        priority = PRIORITY.get(cell("priority").lower().strip(". "))
        training_type = TYPE_FIX.get(cell("type").lower().strip(". "), cell("type") or None)
        institution = cell("institution") or None

        rows.append(
            {
                "name": current,
                "serial": serial or counter,
                "course_title": title,
                "institution": institution,
                "training_dates": cell("date") or None,
                "priority": priority,
                "training_type": training_type,
                "cost": amount,
                "currency": currency,
                # "NSIB" in the institution column is the bureau training its own —
                # the same thing the DG means by "an in-house expert can do it".
                "delivery": "In-house" if institution and institution.strip().upper() in {"NSIB", "OJT"} else "External",
            }
        )
    return rows


def main():
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)
    items = []
    for name in workbook.sheetnames:
        if not re.fullmatch(r"20\d{2}", name):
            continue
        for row in read_year(workbook[name]):
            items.append({"year": int(name), **row})

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf8") as handle:
        json.dump({"items": items}, handle, indent=1, ensure_ascii=False)

    years = sorted({item["year"] for item in items})
    people = sorted({item["name"] for item in items})
    print(json.dumps({"out": OUT, "items": len(items), "years": years, "staff": len(people)}, indent=2))


if __name__ == "__main__":
    main()
