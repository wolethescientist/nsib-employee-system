"""Extract the NSIB Individual Development Plan workbook into data/idp-dataset.json.

Each employee lives on their own sheet with an identical 43-row course grid
(rows 11-53). The grid is the canonical catalogue: row slot + programme type +
course title. Titles are spelled inconsistently across sheets, so the canonical
spelling for each slot is the most common one, with known typos corrected.

Run:  python scripts/extract-workbook.py
"""
import collections
import json
import os
import re

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WORKBOOK = os.path.join(ROOT, "NSIB Individual Development Plan Update(1) (1).xlsx")
OUT = os.path.join(ROOT, "data", "idp-dataset.json")

FIRST_ROW, LAST_ROW = 11, 53
COL = {"no": 2, "type": 3, "title": 4, "priority": 5, "planned": 6, "status": 7, "completed": 8, "unit": 9, "comments": 10}

# Programme types in the order NSIB progresses through them.
TYPE_ORDER = ["Initial", "OJT", "Basic", "Advanced", "Additional", "Specialty", "Recurrent"]
TYPE_FIX = {"Advance": "Advanced", "Sspecialty": "Specialty", "OGT": "OJT"}

# Typos in the winning spelling of a slot, corrected once here.
TITLE_FIX = {
    "Aircrfat type familiarization": "Aircraft Type Familiarization",
    "Cabin Crerw Training( As Applicable)": "Cabin Crew Training (As applicable)",
    "Aviation Maintanance Investigation": "Aviation Maintenance Investigation",
    "Train-the Trainer": "Train-the-Trainer",
    "Fundamental of Accident Investigation": "Fundamentals of Accident Investigation",
    "safety management systems": "Safety Management Systems",
    "family assistance and media relations": "Family Assistance and Media Relations",
    "flight data analysis": "Flight Data Analysis",
    "fires and explosions": "Fires and Explosions",
    "Accident survival aspects": "Accident Survival Aspects",
    "Human Factors investigation": "Human Factors Investigation",
    "Aircraft accident report writing": "Aircraft Accident Report Writing",
    "Aircraft recovery and salvage": "Aircraft Recovery and Salvage",
    "Aircraft Type courses": "Aircraft Type Courses",
    "Site hazard and risk management": "Site Hazard and Risk Management",
    "Accident site drill": "Accident Site Drill",
    "Table top exercises": "Table Top Exercises",
    "Air Traffic Investigation & Analyses": "Air Traffic Investigation & Analyses",
}

PRIORITY = {"high": "P1", "higb": "P1", "medium": "P2", "meddium": "P2", "low": "P3", "r": "R"}
STATUS = {"completed": "Completed", "planned": "Planned", "in progress": "In progress"}

# The workbook has no profession column, but the qualifications line names the
# professional licence each investigator holds, and those markers are
# definitional rather than a guess: AMEL *is* an Aircraft Maintenance Engineer
# Licence, ATC an air traffic controller rating, CPL/ATPL a pilot licence.
# Anything without a licence marker is left blank for an administrator to fill.
# First match wins, so the more specific rating is listed before the general one.
PROFESSION_RULES = [
    (re.compile(r"\bATPL\b|\bCPL\b|\bPPL\b|Commercial Pilot|Pilot Licen", re.I), "Pilot"),
    (re.compile(r"\bATC\b|Air Traffic", re.I), "Air Traffic Controller"),
    (re.compile(r"Avionic", re.I), "Avionics Engineer"),
    (re.compile(r"\bLAME\b|\bAMEL\b|\bAME\b|A ?& ?P|Aircraft Maintenance Engineer", re.I), "Aircraft Maintenance Engineer"),
    (re.compile(r"Dispatcher", re.I), "Flight Dispatcher"),
]


def profession_from(qualifications):
    for pattern, profession in PROFESSION_RULES:
        if pattern.search(qualifications or ""):
            return profession
    return None


def norm(value):
    return re.sub(r"\s+", " ", str(value if value is not None else "")).strip()


def year(value):
    """Excel stores completion as a bare year; anything else is unusable."""
    text = norm(value)
    match = re.search(r"(19|20)\d{2}", text)
    if not match:
        return None
    parsed = int(match.group(0))
    return parsed if 1950 <= parsed <= 2100 else None


# The licence is written a different way on nearly every sheet: `( License 5469)`,
# `(License: 3846)`, `(License 2673)/ 5079`, `License  6080`, `( 3464)`,
# `(Lince : 679)`, `(ATPL License no 4938A)`, even the unbalanced `( License 8042( A)`.
# Matching the word is hopeless — the dependable signal is a 3-6 digit number that
# appears after an opening bracket or a licence-ish word, always at the end of the
# name. So find where that trailing block starts, cut from there, and keep the digits.
LICENCE_START = re.compile(r"[(\[]|\b(?:lic\w*|linc\w*|amel|atpl)\b", re.I)
LICENCE_NUMBER = re.compile(r"\d{3,6}[A-Za-z]?")


def split_license(raw_name):
    """`Engr. Aliyu Umar ( License 5469)` -> ('Engr. Aliyu Umar', '5469')."""
    name = norm(raw_name)
    for match in LICENCE_START.finditer(name):
        tail = name[match.start():]
        numbers = LICENCE_NUMBER.findall(tail)
        if not numbers:
            continue
        return norm(name[: match.start()].strip(" ,;:/-")), " / ".join(numbers)
    return norm(name.strip(" ,;:/-")), None


def email_for(name, taken):
    parts = [re.sub(r"[^a-z]", "", part.lower()) for part in name.split()]
    parts = [part for part in parts if part and part not in {"engr", "capt", "fo", "mr", "mrs", "ms", "dr"}]
    if not parts:
        parts = ["staff"]
    base = f"{parts[0]}.{parts[-1]}" if len(parts) > 1 else parts[0]
    candidate, suffix = base, 2
    while candidate in taken:
        candidate, suffix = f"{base}{suffix}", suffix + 1
    taken.add(candidate)
    return f"{candidate}@nsib.gov.ng"


def qualifications(cell_value):
    text = norm(cell_value)
    text = re.sub(r"^QUALIFICATIONS\s*WITH\s*DATES\s*[;:,]?\s*", "", text, flags=re.I)
    return text or None


def experience_years(value):
    match = re.search(r"\d+", norm(value))
    return int(match.group(0)) if match else None


def main():
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)
    sheets = [sheet for sheet in workbook.worksheets if sheet.title.strip() != "Sheet1"]

    # Canonical catalogue: most common (type, title) per row slot.
    per_slot = collections.defaultdict(collections.Counter)
    for sheet in sheets:
        for row in range(FIRST_ROW, LAST_ROW + 1):
            title = norm(sheet.cell(row, COL["title"]).value)
            if title:
                per_slot[row][(norm(sheet.cell(row, COL["type"]).value), title)] += 1

    courses = []
    for index, row in enumerate(sorted(per_slot)):
        (raw_type, raw_title), _ = per_slot[row].most_common(1)[0]
        programme_type = TYPE_FIX.get(raw_type, raw_type)
        assert programme_type in TYPE_ORDER, f"unknown programme type {raw_type!r} at row {row}"
        courses.append({
            "slot": row,
            "programme_type": programme_type,
            "name": TITLE_FIX.get(raw_title, raw_title),
            "sort_order": index + 1,
        })

    employees, records, taken = [], [], set()
    for sheet in sheets:
        name, licence = split_license(sheet["D3"].value)
        if not name:
            continue
        key = sheet.title.strip()
        quals = qualifications(sheet["B9"].value)
        employees.append({
            "key": key,
            "name": name,
            "license": licence,
            "designation": norm(sheet["D4"].value) or None,
            "division": norm(sheet["D5"].value) or None,
            "department": norm(sheet["D6"].value) or None,
            "profession": profession_from(quals),
            "training_profile": norm(sheet["D7"].value) or None,
            "years_experience": experience_years(sheet["D8"].value),
            "qualifications": quals,
            "email": email_for(name, taken),
        })
        for course in courses:
            row = course["slot"]
            unit = norm(sheet.cell(row, COL["unit"]).value).lower()
            status = STATUS.get(norm(sheet.cell(row, COL["status"]).value).lower())
            completed_year = year(sheet.cell(row, COL["completed"]).value)
            if completed_year and not status:
                status = "Completed"
            records.append({
                "employee_key": key,
                "course_slot": row,
                # Blank operations-unit cells default to applicable, matching the sheet.
                "applicable": unit != "not applicable",
                "priority": PRIORITY.get(norm(sheet.cell(row, COL["priority"]).value).lower().replace(" ", "")),
                "planned_year": year(sheet.cell(row, COL["planned"]).value),
                "status": status or "Not started",
                "completed_year": completed_year,
                "comments": norm(sheet.cell(row, COL["comments"]).value) or None,
            })

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    payload = {"programmeTypes": TYPE_ORDER, "courses": courses, "employees": employees, "records": records}
    with open(OUT, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, ensure_ascii=False)

    professions = collections.Counter(employee["profession"] or "(not recorded)" for employee in employees)
    print(json.dumps({
        "out": OUT,
        "courses": len(courses),
        "employees": len(employees),
        "records": len(records),
        "completed": sum(1 for r in records if r["status"] == "Completed"),
        "notApplicable": sum(1 for r in records if not r["applicable"]),
        "professionsFromLicence": dict(professions.most_common()),
    }, indent=2))


if __name__ == "__main__":
    main()
