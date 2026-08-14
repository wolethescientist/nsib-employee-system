import openpyxl, collections

p = r'C:\Users\david\Downloads\NSIB Individual Development Plan Update(1) (1).xlsx'
w = openpyxl.load_workbook(p, read_only=True, data_only=False)
print('SHEET_COUNT', len(w.sheetnames))
print('SHEET_NAMES', ', '.join(w.sheetnames))
statuses = collections.Counter()
phases = collections.Counter()
employees = []
completed = planned = links = 0

for s in w.worksheets:
    vals = [[c.value for c in row] for row in s.iter_rows(min_row=1, max_row=60, max_col=10)]
    def field(row):
        return str(vals[row][3]).strip() if row < len(vals) and len(vals[row]) > 3 and vals[row][3] else ''
    employees.append((s.title, field(2), field(3), field(4)))
    for r in vals[10:53]:
        if len(r) > 3 and r[3]:
            phases[str(r[2]).strip()] += 1
            if r[6]: statuses[str(r[6]).strip()] += 1
            if str(r[6]).strip().lower() == 'completed': completed += 1
            if str(r[6]).strip().lower() == 'planned': planned += 1
            if r[9] and ('HYPERLINK' in str(r[9]).upper() or 'http' in str(r[9]).lower()): links += 1

print('STATUS', dict(statuses))
print('PHASES', dict(phases))
print('COMPLETED', completed, 'PLANNED', planned, 'DOCUMENT_LINK_CELLS', links)
print('FIRST_15_EMPLOYEES')
for x in employees[:15]: print(x)
